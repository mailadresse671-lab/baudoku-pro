import os
import shutil
import time
import json
from datetime import datetime
from PIL import Image
import google.generativeai as genai
import openpyxl 

# --- KONFIGURATION ---
API_KEY = "AIzaSyAgDpAkTJlbqWADqqGrvZFFZp76tznY5C0"  
BASIS_PFAD = "Bau_Projekte"

def setup_gemini(lv_path):
    if not API_KEY or "HIER" in API_KEY:
        return None
    genai.configure(api_key=API_KEY)
    
    # LV Einlesen
    lv_content = "Kein LV vorhanden."
    if os.path.exists(lv_path):
        files = [f for f in os.listdir(lv_path) if f.lower().endswith('.txt')]
        if files:
            try:
                with open(os.path.join(lv_path, files[0]), 'r', encoding='utf-8') as f:
                    lv_content = f.read()
            except: pass

    system_prompt = f"""
    Du bist Bauleiter (Tiefbau). Erstelle Daten für einen Excel-Bautagesbericht.
    LV-Kontext: {lv_content[:50000]}
    Inventar: Takeuchi 6t/16t, Rüttelplatte, Grabenstampfer, Stromaggregat.
    
    ANTWORTE NUR ALS JSON:
    {{
      "wetter_morgens": "String",
      "wetter_abends": "String",
      "temp_min": "String",
      "temp_max": "String",
      "personal_aufsicht": 1,
      "personal_facharbeiter": 2,
      "personal_maschinist": 1,
      "beschreibung_arbeiten": "Text",
      "geraete_liste": "Text",
      "material_liste": "Text",
      "sonstiges": "Text"
    }}
    """
    
    # Hier ist der Fix: Wir geben keine fixen Modellnamen vor, sondern testen im Main-Loop
    return system_prompt

def fill_excel_template(template_path, output_path, data, date_str, project_name):
    try:
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active 
        
        ws['C2'] = "Auftraggeber" 
        ws['C3'] = project_name
        ws['C4'] = date_str 
        
        ws['B8'] = data.get('wetter_morgens', '')
        ws['B9'] = data.get('wetter_abends', '')
        ws['H8'] = data.get('temp_min', '')
        ws['H9'] = data.get('temp_max', '')
        
        p1 = data.get('personal_aufsicht', 0)
        p2 = data.get('personal_facharbeiter', 0)
        p3 = data.get('personal_maschinist', 0)
        ws['C11'] = p1
        ws['C12'] = p2
        ws['C13'] = p3
        ws['C15'] = p1 + p2 + p3
        
        ws['B17'] = data.get('beschreibung_arbeiten', '')
        ws['B32'] = data.get('geraete_liste', '')
        ws['B37'] = data.get('material_liste', '')
        ws['B41'] = data.get('sonstiges', '')

        wb.save(output_path)
        return True
    except Exception as e:
        print(f"❌ Excel Fehler: {e}")
        return False

def get_date_taken(path):
    """Verbesserte Datums-Erkennung"""
    try:
        # Versuch 1: Echtes EXIF Aufnahmedatum
        img = Image.open(path)
        exif = img.getexif()
        if exif and 36867 in exif:
            return exif[36867] # Format: YYYY:MM:DD HH:MM:SS
    except:
        pass
    
    # Versuch 2: Dateisystem Erstellungsdatum (Nicht Änderungsdatum!)
    return time.strftime('%Y:%m:%d %H:%M:%S', time.gmtime(os.path.getctime(path)))

def try_generate_content(file_paths, prompt_text, sys_instruct):
    """Probiert verschiedene Modelle durch, falls eines 404 gibt"""
    # Liste der Modelle, die wir probieren (von neu nach alt)
    models_to_try = ["gemini-1.5-flash", "gemini-1.5-flash-latest", "gemini-1.5-pro", "gemini-pro-vision"]
    
    uploaded_files = []
    for p in file_paths:
        uploaded_files.append(genai.upload_file(p))

    for model_name in models_to_try:
        try:
            model = genai.GenerativeModel(model_name=model_name, system_instruction=sys_instruct, generation_config={"response_mime_type": "application/json"})
            response = model.generate_content([prompt_text, *uploaded_files])
            return response
        except Exception as e:
            if "404" in str(e):
                print(f"⚠️ Modell {model_name} nicht gefunden, versuche nächstes...")
                continue
            else:
                print(f"❌ Anderer Fehler bei {model_name}: {e}")
                return None
    return None

def main():
    print("--- 🏗️ EXCEL-ROBOTER V2 (FIX) ---")
    
    if not os.path.exists(BASIS_PFAD):
        print("❌ Keine Projekte gefunden."); input(); return

    projekte = [d for d in os.listdir(BASIS_PFAD) if os.path.isdir(os.path.join(BASIS_PFAD, d))]
    print("\nVerfügbare Projekte:")
    for i, p in enumerate(projekte): print(f"[{i+1}] {p}")
    
    try:
        wahl = int(input("\nProjekt wählen: ")) - 1
        curr_proj = projekte[wahl]
    except: return

    root = os.path.join(BASIS_PFAD, curr_proj)
    src_folder = os.path.join(root, "Eingang_Fotos")
    dest_base = os.path.join(root, r"1.4 Berichte\1.4.1 Tagesberichte\Fertig")
    template_folder = os.path.join(root, r"1.4 Berichte\1.4.1 Tagesberichte\Vorlagen")
    lv_folder = os.path.join(root, "Projekt_Infos")

    # Vorlage suchen
    template_file = None
    if os.path.exists(template_folder):
        files = [f for f in os.listdir(template_folder) if f.endswith('.xlsx')]
        if files: template_file = os.path.join(template_folder, files[0])
    
    if not template_file:
        print("⚠️ KEINE EXCEL-VORLAGE GEFUNDEN!"); input(); return

    # Prompt vorbereiten
    sys_prompt = setup_gemini(lv_folder)
    if not sys_prompt: print("❌ API Key fehlt!"); input(); return

    # Sortieren
    days_todo = {} 

    print(f"\n📂 Scanne Bilder in {curr_proj}...")
    for file in os.listdir(src_folder):
        if file.lower().endswith(('jpg', 'jpeg', 'png')):
            full_src = os.path.join(src_folder, file)
            try:
                date_str = get_date_taken(full_src)
                dt = datetime.strptime(date_str, '%Y:%m:%d %H:%M:%S')
                folder_date = dt.strftime('%Y-%m-%d')
                
                if folder_date not in days_todo: days_todo[folder_date] = []
                days_todo[folder_date].append(full_src)
            except: pass

    print(f"\n🤖 Gefunden: {len(days_todo)} Arbeitstage.")
    
    for date_key in sorted(days_todo.keys()):
        day_folder = os.path.join(dest_base, date_key)
        if not os.path.exists(day_folder): os.makedirs(day_folder)
        
        excel_out = os.path.join(day_folder, f"Bautagesbericht_{date_key}.xlsx")
        
        if os.path.exists(excel_out):
            print(f"Skipping {date_key} (schon fertig).")
            continue

        print(f"⏳ {date_key}: Fülle Excel...")
        imgs = days_todo[date_key][:10]
        
        # Bilder kopieren
        for img_path in imgs:
            try: shutil.copy2(img_path, day_folder)
            except: pass

        # KI Generierung mit Fallback
        response = try_generate_content(imgs, f"Bericht für {date_key}", sys_prompt)
        
        if response and response.text:
            try:
                json_data = json.loads(response.text)
                success = fill_excel_template(template_file, excel_out, json_data, date_key, curr_proj)
                if success: print(f"✅ Excel gespeichert!")
            except Exception as e:
                print(f"❌ Fehler beim Speichern: {e}")
        else:
            print("❌ KI konnte keine Daten generieren (404 oder anderer Fehler).")
            
        time.sleep(4)

    print("\n✅ FERTIG!")
    input()

if __name__ == "__main__":
    main()