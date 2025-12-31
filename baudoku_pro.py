import os, shutil, time, json, logging, warnings, re, textwrap
from datetime import datetime
from PIL import Image
import openpyxl
from openpyxl.styles import Alignment
import google.generativeai as genai
import pypdf

warnings.filterwarnings("ignore")

# =========================
# KONFIGURATION
# =========================
API_KEY = "AIzaSyAgDpAkTJlbqWADqqGrvZFFZp76tznY5C0"
BASIS_PFAD = "Bau_Projekte"

MAX_TAGE_TEST = 5
MAX_BILDER_PRO_TAG = 3
LOESCHEN_AKTIV = False
ZEICHEN_PRO_ZEILE = 90
DRY_RUN = False

PROJEKT_AUFTRAGGEBER = "TransnetBW Erw. Umspannwerk"
PROJEKT_BAUSTELLE = "810325005 Grünkraut"
PROJEKT_BEARBEITER = "Beham"

COUNTER_FILE = os.path.join(BASIS_PFAD, "bericht_zaehler.json")

# =========================
# LOGGING
# =========================
def setup_logger(project):
    log_path = os.path.join(BASIS_PFAD, project, "baudoku.log")
    logging.basicConfig(
        filename=log_path,
        level=logging.INFO,
        format="%(asctime)s | %(levelname)s | %(message)s"
    )
    logging.info("===== PROGRAMMSTART =====")

# =========================
# STATUS (RESUME)
# =========================
def status_file(project):
    return os.path.join(BASIS_PFAD, project, "status.json")

def load_status(project):
    try:
        with open(status_file(project), "r") as f:
            return json.load(f)
    except:
        return {}

def save_status(project, status):
    with open(status_file(project), "w") as f:
        json.dump(status, f, indent=2)

# =========================
# BERICHTSNUMMERN (SAFE)
# =========================
def load_counter():
    try:
        with open(COUNTER_FILE, "r") as f:
            return json.load(f)
    except:
        return {}

def save_counter(data):
    with open(COUNTER_FILE, "w") as f:
        json.dump(data, f, indent=2)

def peek_next_report_nr(project):
    return load_counter().get(project, 0) + 1

def commit_report_nr(project):
    data = load_counter()
    data[project] = data.get(project, 0) + 1
    save_counter(data)
    return data[project]

# =========================
# HILFSFUNKTIONEN
# =========================
def safe_json_load(text):
    try:
        s = text[text.find("{"):text.rfind("}") + 1]
        return json.loads(s)
    except:
        return None

def plausibility_check(d):
    w = []
    # Typumwandlung sicherstellen
    try: t_min = float(str(d.get("temp_min",0)).replace(',','.'))
    except: t_min = 0
    try: t_max = float(str(d.get("temp_max",0)).replace(',','.'))
    except: t_max = 0
    
    if t_min > t_max:
        w.append(f"Temperatur unlogisch: Min {t_min} > Max {t_max}")
    
    fach = int(d.get("personal_facharbeiter", 0))
    if fach > 20:
        w.append(f"Ungewöhnlich viele Facharbeiter: {fach}")
        
    if not d.get("beschreibung_arbeiten"):
        w.append("Keine Arbeiten erkannt")
    return w

def get_date_taken(path, filename):
    try:
        exif = Image.open(path).getexif()
        if 36867 in exif:
            return exif[36867]
    except:
        pass
    m = re.search(r'(20\d{2})[-_]?(\d{2})[-_]?(\d{2})', filename)
    if m:
        return f"{m[1]}:{m[2]}:{m[3]} 12:00:00"
    return datetime.fromtimestamp(os.path.getmtime(path)).strftime('%Y:%m:%d %H:%M:%S')

# =========================
# GEMINI PROMPT
# =========================
def build_prompt(lv_text):
    return f"""
Du bist ein erfahrener Bauleiter im Tief-, Straßen- und Hochbau.
Erstelle einen realistischen Bautagesbericht basierend auf den Fotos.

REGELN:
- Erfinde nichts.
- Wenn keine LV-Position passt: "keine eindeutige LV-Position gefunden".
- Antworte AUSSCHLIESSLICH als JSON.

LEISTUNGSVERZEICHNIS (Auszug):
{lv_text[:100000]} 

JSON STRUKTUR:
{{
  "wetter_vormittag": "Text",
  "wetter_nachmittag": "Text",
  "temp_min": 0,
  "temp_max": 0,
  "personal_aufsicht": 0,
  "personal_facharbeiter": 0,
  "personal_maschinist": 0,
  "beschreibung_arbeiten": ["Zeile 1", "Zeile 2"],
  "geraete_liste": ["Gerät 1", "Gerät 2"],
  "material_liste": ["Material 1", "Material 2"],
  "sonstiges": []
}}
"""

# =========================
# GEMINI CALL
# =========================
def generate_report(images, prompt):
    genai.configure(api_key=API_KEY)
    
    # Retry Logik für Quota Errors
    retries = 3
    delay = 60
    
    # Modelle durchprobieren
    models = ["gemini-2.0-flash", "gemini-1.5-flash"]
    
    for m in models:
        for i in range(retries):
            try:
                model = genai.GenerativeModel(
                    m,
                    generation_config={"response_mime_type": "application/json"}
                )
                uploads = [genai.upload_file(i) for i in images]
                return model.generate_content([prompt, *uploads])
            except Exception as e:
                if "429" in str(e):
                    print(f"⚠️ Quota Limit ({m}). Warte {delay}s...")
                    time.sleep(delay)
                    delay += 30
                elif "404" in str(e):
                    break # Modell nicht da, nächstes
                else:
                    logging.error(f"API Fehler: {e}")
                    break
    return None

# =========================
# EXCEL
# =========================
def fill_excel(path, data, date_str, nr):
    wb = openpyxl.load_workbook(path, keep_vba=True)
    
    # Tab wählen
    wochentage = ["Montag","Dienstag","Mittwoch","Donnerstag","Freitag","Samstag","Sonntag"]
    tag_name = wochentage[datetime.strptime(date_str,'%d.%m.%Y').weekday()]
    
    if tag_name in wb.sheetnames:
        ws = wb[tag_name]
    else:
        ws = wb.active

    def w(cell,v): 
        try:
            ws[cell].value=v
            ws[cell].alignment=Alignment(wrap_text=True, vertical='top', horizontal='left')
        except: pass

    # Hilfsfunktion für Listen (Zeile für Zeile)
    def w_list(start_cell, items):
        if not items: return
        # Koordinaten
        col_char = "".join([c for c in start_cell if c.isalpha()])
        start_row = int("".join([c for c in start_cell if c.isdigit()]))
        from openpyxl.utils import column_index_from_string
        col_idx = column_index_from_string(col_char)
        
        curr = start_row
        for item in items:
            # Umbruch simulieren
            lines = textwrap.wrap(str(item), width=ZEICHEN_PRO_ZEILE)
            for line in lines:
                c = ws.cell(row=curr, column=col_idx)
                c.value = line
                c.alignment = Alignment(wrap_text=False, vertical='bottom', horizontal='left')
                curr += 1

    # Header
    w("F1",nr); w("D1",nr)
    w("C2",PROJEKT_AUFTRAGGEBER)
    w("C3",PROJEKT_BAUSTELLE)
    w("C4",date_str)
    w("C6",PROJEKT_BEARBEITER)

    # Wetter
    w("B8",data.get("wetter_vormittag", ""))
    w("B9",data.get("wetter_nachmittag", ""))
    w("H8",data.get("temp_min", ""))
    w("H9",data.get("temp_max", ""))

    # Personal
    p1 = int(data.get("personal_aufsicht", 0))
    p2 = int(data.get("personal_facharbeiter", 0))
    p3 = int(data.get("personal_maschinist", 0))
    
    w("C11",p1)
    w("C12",p2)
    w("C13",p3)
    w("C15", p1+p2+p3)

    # Listeninhalte
    w_list("B17", data.get("beschreibung_arbeiten", []))
    w_list("B32", data.get("geraete_liste", []))
    w_list("B37", data.get("material_liste", []))
    w_list("B41", data.get("sonstiges", [])) # Sonstiges ist bei B41

    if not DRY_RUN:
        wb.save(path)

# =========================
# MAIN
# =========================
def main():
    print("🏗️ BAUDOKU START V33 (FINAL)")
    
    if not os.path.exists(BASIS_PFAD):
        print("❌ Keine Projekte gefunden.")
        return

    projects = [p for p in os.listdir(BASIS_PFAD) if os.path.isdir(os.path.join(BASIS_PFAD,p))]
    for i,p in enumerate(projects): print(f"[{i+1}] {p}")
    
    try:
        curr_proj = projects[int(input("Projekt: "))-1]
    except: return

    setup_logger(curr_proj)
    status = load_status(curr_proj)

    src = os.path.join(BASIS_PFAD,curr_proj,"Eingang_Fotos")
    tpl_dir = os.path.join(BASIS_PFAD,curr_proj,"1.4 Berichte","1.4.1 Tagesberichte","Vorlagen")
    out_dir = os.path.join(BASIS_PFAD,curr_proj,"1.4 Berichte","1.4.1 Tagesberichte","Fertig")
    
    lv_path = os.path.join(BASIS_PFAD,curr_proj,"Projekt_Infos")

    # LV Laden
    lv_text = ""
    if os.path.exists(lv_path):
        for f in os.listdir(lv_path):
            if f.endswith(".txt"):
                try: 
                    with open(os.path.join(lv_path,f),'r',encoding='utf-8') as file: lv_text+=file.read()
                except: pass
            elif f.endswith(".pdf"):
                try:
                    r = pypdf.PdfReader(os.path.join(lv_path,f))
                    for p in r.pages: lv_text+=p.extract_text()
                except: pass
    
    # Vorlage suchen
    tpl_file = None
    if os.path.exists(tpl_dir):
        for f in os.listdir(tpl_dir):
            if f.endswith((".xlsx",".xlsm")) and not f.startswith("~$"):
                if "tages" in f.lower() or "bau" in f.lower():
                    tpl_file = os.path.join(tpl_dir,f)
                    break
    
    if not tpl_file:
        print("⚠️ Vorlage nicht gefunden!")
        return

    # Bilder scannen
    days={}
    for r,_,fs in os.walk(src):
        for f in fs:
            if f.lower().endswith(("jpg","png","jpeg")):
                path = os.path.join(r,f)
                d = get_date_taken(path,f)[:10] # YYYY:MM:DD
                days.setdefault(d,[]).append(path)

    print(f"🤖 {len(days)} Tage gefunden.")
    
    count = 0
    # Sortiert abarbeiten
    for date_key in sorted(days):
        if count >= MAX_TAGE_TEST: 
            print("🛑 Test-Limit erreicht.")
            break
        
        # Datum Formatierung
        dt_obj = datetime.strptime(date_key,"%Y:%m:%d")
        date_fmt = dt_obj.strftime("%d.%m.%Y") # 14.07.2025
        
        # Check Status
        if status.get(date_key) == "done":
            # print(f"Skipping {date_fmt}")
            continue

        count += 1
        logging.info(f"{date_fmt}: START")
        print(f"⏳ {date_fmt}: Generiere...")

        try:
            # 1. Nummer holen
            preview_nr = peek_next_report_nr(curr_proj)
            
            # 2. KI Fragen
            imgs = days[date_key][:MAX_BILDER_PRO_TAG]
            resp = generate_report(imgs, build_prompt(lv_text))
            
            if not resp: raise Exception("KI lieferte keine Antwort")
            
            data = safe_json_load(resp.text)
            if not data: raise ValueError("JSON ungültig")

            # 3. Plausibilitätscheck
            warns = plausibility_check(data)
            if warns: 
                data["sonstiges"].extend([f"⚠️ {w}" for w in warns])
                print(f"   ⚠️ Warnungen: {warns}")

            # 4. Excel vorbereiten
            kw = dt_obj.isocalendar()[1]
            year = dt_obj.year
            
            # Ordner KW29_2025
            kw_dir = os.path.join(out_dir, f"KW{kw}_{year}")
            os.makedirs(kw_dir, exist_ok=True)
            
            # Excel Datei
            ext = os.path.splitext(tpl_file)[1]
            excel_name = f"Bautagesbericht_KW{kw}{ext}"
            excel_path = os.path.join(kw_dir, excel_name)
            
            # Bilder Ordner Tag
            img_dir = os.path.join(kw_dir, date_fmt, "Bilder")
            os.makedirs(img_dir, exist_ok=True)

            # Wenn Excel nicht da -> Kopieren
            if not os.path.exists(excel_path):
                shutil.copy2(tpl_file, excel_path)

            # 5. Bilder verschieben/kopieren
            for img_p in days[date_key]:
                fname = os.path.basename(img_p)
                tgt = os.path.join(img_dir, fname)
                try:
                    if LOESCHEN_AKTIV: shutil.move(img_p, tgt)
                    else: shutil.copy2(img_p, tgt)
                except: pass

            # 6. Excel Füllen
            fill_excel(excel_path, data, date_fmt, preview_nr)
            
            # 7. Finalisieren
            final_nr = commit_report_nr(curr_proj)
            status[date_key] = "done"
            logging.info(f"{date_fmt}: OK (Nr {final_nr})")
            print(f"   ✅ Fertig! Nr. {final_nr}")

        except Exception as e:
            status[date_key] = "failed"
            logging.error(f"{date_fmt}: FEHLER {e}")
            print(f"   ❌ Fehler: {e}")

        save_status(curr_proj, status)
        print("   💤 Pause...")
        time.sleep(45)

    print("✅ FERTIG")

if __name__ == "__main__":
    main()